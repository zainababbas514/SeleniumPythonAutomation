import os
import inspect
import logging
import random
import openpyxl
from datetime import datetime
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class BaseClass:

    @staticmethod
    def get_logger():
        log_path = f"logs/global_sqa_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
        os.makedirs(os.path.dirname(log_path), exist_ok=True)

        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        logger.setLevel(logging.DEBUG)

        if not logger.handlers:
            file_handler = logging.FileHandler(log_path)
            file_handler.setLevel(logging.DEBUG)

            console_handler = logging.StreamHandler()
            console_handler.setLevel(logging.DEBUG)

            formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
            file_handler.setFormatter(formatter)
            console_handler.setFormatter(formatter)

            logger.addHandler(file_handler)
            logger.addHandler(console_handler)

        logger.propagate = True
        return logger

    def wait_for_url_contains_text(self, text, timeout = 10):
        WebDriverWait(self.driver, timeout).until(EC.url_contains(text))

    def wait_for_element_visibility(self, element, timeout = 10):
        return WebDriverWait(self.driver, timeout).until(EC.visibility_of_element_located(element))

    def wait_for_all_element_visibility(self, element, timeout = 10):
        return WebDriverWait(self.driver, timeout).until(EC.visibility_of_all_elements_located(element))

    def generate_random_number(self, range_max):
        if range_max < 1:
            raise ValueError("The specified range is too small")
        return random.randint(0, range_max-1)

    def wait_for_page_load(self, url_text, locator):
        self.wait_for_url_contains_text(url_text)
        self.wait_for_element_visibility(locator)

    def open_tab(self, locator):
        tab = (self.driver.find_element(*locator))
        if "resp-tab-active" not in tab.get_attribute('class'):
            self.driver.execute_script("arguments[0].scrollIntoView(true)", tab)
            self.driver.execute_script("arguments[0].click()", tab)

    def switch_to_frame(self, frame):
        if isinstance(frame, str):
            self.driver.switch_to.frame(frame)
        else:
            frame = self.driver.find_element(*frame)
            self.driver.switch_to.frame(frame)

    @staticmethod
    def get_test_data(test_case_name, sheet):
        all_data = []
        workbook = openpyxl.load_workbook("testData/TestData.xlsx")
        sheet = workbook[sheet]
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(i, 1).value == test_case_name:
                data = {}
                for j in range(2, sheet.max_column + 1):
                    data[sheet.cell(1, j).value] = sheet.cell(i, j).value
                all_data.append(data)
        return all_data

    def get_alert(self, timeout = 10):
        return WebDriverWait(self.driver, timeout).until(EC.alert_is_present())

    def accept_alert(self):
        alert = self.get_alert()
        text = alert.text
        alert.accept()
        return text

    def wait_for_text_present(self, element, text, timeout = 10):
        return WebDriverWait(self.driver, timeout).until(EC.text_to_be_present_in_element(element, text))

    def wait_for_element_invisibility(self, element, timeout = 10):
        return WebDriverWait(self.driver, timeout).until(EC.invisibility_of_element_located(element))





